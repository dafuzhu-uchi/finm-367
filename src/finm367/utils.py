import polars as pl
from openpyxl import load_workbook
from pathlib import Path
from typing import List, Dict
from datetime import datetime
import calendar
import math

def load_path(file_name):
    data_path = Path.cwd().parents[1] / "data"
    file_path = data_path / file_name
    return file_path

def print_sheetname(FILE_PATH: Path) -> List:  #@save
    wb = load_workbook(FILE_PATH, read_only=True)
    result = wb.sheetnames
    wb.close()
    return result

def to_frame(series):
    return series.to_frame() if isinstance(series, pl.Series) else series

def only_numeric(data: pl.DataFrame) -> pl.DataFrame:
    numeric = [pl.Float64, pl.Float32, pl.Int64, pl.Int32, pl.Int16, pl.Int8]
    cols = [col for col in data.columns if data[col].dtype in numeric]
    return data.select(pl.col(cols))


def get_tickers(data: pl.DataFrame) -> List[str]:
    return only_numeric(data).columns


def concat_with_labels(
    dfs: Dict[str, pl.DataFrame] | List[pl.DataFrame],
    label_col: str = "statistic",
    labels: List[str] | None = None,
    label_position: str = "first"
) -> pl.DataFrame:
    """
    Concatenate DataFrames vertically and add a label column.
    
    Parameters
    ----------
    dfs : Dict[str, pl.DataFrame] or List[pl.DataFrame]
        DataFrames to concatenate. If dict, keys are used as labels.
        If list, labels must be provided separately.
    label_col : str, default "statistic"
        Name of the label column to add.
    labels : List[str] or None, default None
        Labels for each DataFrame. Only needed if dfs is a list.
        Ignored if dfs is a dict.
    label_position : str, default "first"
        Position of label column: "first" or "last".
    
    Returns
    -------
    pl.DataFrame
        Concatenated DataFrame with label column.
    """
    # Convert to dict if list is provided
    if isinstance(dfs, list):
        if labels is None:
            raise ValueError("labels must be provided when dfs is a list")
        if len(dfs) != len(labels):
            raise ValueError(f"Number of DataFrames ({len(dfs)}) must match number of labels ({len(labels)})")
        dfs = dict(zip(labels, dfs))
    
    # Add label column to each DataFrame
    labeled_dfs = [
        df.with_columns(pl.lit(label).alias(label_col))
        for label, df in dfs.items()
    ]
    
    # Concatenate vertically
    result = pl.concat(labeled_dfs, how="vertical")
    
    # Reorder columns based on label_position
    if label_position == "first":
        result = result.select([label_col, pl.all().exclude(label_col)])
    elif label_position == "last":
        result = result.select([pl.all().exclude(label_col), label_col])
    else:
        raise ValueError(f"label_position must be 'first' or 'last', got '{label_position}'")
    
    return result

def broadcast_rows(
    df: pl.DataFrame,
    scaler: pl.DataFrame
) -> pl.DataFrame:
    """
    Broadcast rows according to scaler.

    Parameters
    ----------
    df : pl.DataFrame, shape (n_rows, n_cols)
        DataFrame to broadcast.
    scaler : pl.DataFrame, shape (1, n_cols)
        Scaler DataFrame.

    Returns
    -------
    broadcasted dataframe, shape (n_rows, n_cols)
    """
    if df.columns != scaler.columns:
        raise ValueError("DataFrames must have same columns")
    return pl.DataFrame([df[col] * scaler[col] for col in df.columns])


def parse_date(date_str: str, is_end: bool = False) -> datetime:
    """Parse date string with flexible format, defaulting to first day."""
    if date_str is None:
        return None

    # Try different formats
    for fmt in ["%Y-%m-%d", "%Y-%m", "%Y"]:
        try:
            dt = datetime.strptime(date_str, fmt)
            if fmt == "%Y-%m" and is_end:
                last_day = calendar.monthrange(dt.year, dt.month)[1]
                return dt.replace(day=last_day)
            elif fmt == "%Y" and is_end:
                return dt.replace(month=12, day=31)
            return dt
        except ValueError:
            continue

    raise ValueError(f"Date string '{date_str}' doesn't match any supported format (%Y-%m-%d, %Y-%m, or %Y)")


def get_period(
        data: pl.DataFrame,
        start_str: str = None,
        end_str: str = None
) -> pl.DataFrame:
    """
    Args:
        - data: pl.DataFrame
        - start_str: Date string in format %Y-%m-%d, %Y-%m, or %Y
        - end_str: Date string in format %Y-%m-%d, %Y-%m, or %Y
    """
    start_day = parse_date(start_str)
    end_day = parse_date(end_str, is_end=True)

    if start_day is None and end_day is None:
        tb = data
    elif start_day and end_day:
        tb = data.filter(pl.col("date").is_between(start_day, end_day))
    elif start_day:
        tb = data.filter(pl.col("date").ge(start_day))
    elif end_day:
        tb = data.filter(pl.col("date").le(end_day))

    return tb


def log_series(series: pl.Series) -> pl.Series:
    return (series + 1).log(base=math.exp(1))


def log_dataframe(df: pl.DataFrame) -> pl.DataFrame:
    return df.with_columns([
        log_series(pl.col(col)).alias(col)
        for col in df.columns if df.schema[col] == pl.Float64
    ])